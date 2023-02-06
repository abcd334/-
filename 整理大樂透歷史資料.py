import docx
import openpyxl
from docx import Document
import pandas as pd

wb = openpyxl.load_workbook('C:\\Users\Arlen\Desktop\Python\樂透開獎\大樂透.xlsx')
ws = wb['工作表1']
k=2
for year_number in range(96,113):
    path = "C:\\Users\Arlen\Desktop\Python\樂透開獎\開獎號碼\\" + str(year_number) + "年度大樂透開獎號碼表.docx"

    document = Document(path)  # 讀入檔案
    tables = document.tables
    for table in tables:
        for i in range(1, len(table.rows)):
            if table.cell(i, 0).text!="":
                if table.cell(i, 1).text!="":
                    open_month = table.cell(i, 1).text

                open_date= str(year_number+1911) +"/" + str(open_month) + "/" + str(table.cell(i, 2).text)

                ws.cell(k, 1, open_date)
                for j in range(2,9):
                    ws.cell(k,j,int(table.cell(i, 8+j).text))
                k+=1
                wb.save('大樂透.xlsx')

                print(open_date)
