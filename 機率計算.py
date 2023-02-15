import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('C:\\Users\Arlen\Desktop\GitHub\win-the-lottery\大樂透開獎號碼-新版.xlsx')
ws = wb['工作表1']
wr = wb['機率']
k = ws.max_row

data=pd.DataFrame()
data_rate=pd.DataFrame()

for i in range(2,k+1):

    N1 = ws.cell(i, 1).value
    N2 = ws.cell(i, 5).value
    N3 = ws.cell(i, 6).value
    N4 = ws.cell(i, 7).value
    N5 = ws.cell(i, 8).value
    N6 = ws.cell(i, 9).value
    N7 = ws.cell(i, 10).value
    N8 = ws.cell(i, 11).value
    N9 = ws.cell(i, 12).value

    data=data.append({
                    '期別':N1,'獎號1':N2,'獎號2':N3,'獎號3':N4,
                    '獎號4':N5,'獎號5':N6,'獎號6':N7,'特別號':N8,'頭獎數量':N9},ignore_index=True)
    if i>100:
        data_rate = data_rate.append({
            '期別': N1,
            '獎號1': round(((data['獎號1']==N2).sum())/(6*(i-1))*100,2),
            '獎號2': round(((data['獎號2']==N3).sum())/(6*(i-1))*100,2),
            '獎號3': round(((data['獎號3']==N4).sum())/(6*(i-1))*100,2),
            '獎號4': round(((data['獎號4']==N5).sum())/(6*(i-1))*100,2),
            '獎號5': round(((data['獎號5']==N6).sum())/(6*(i-1))*100,2),
            '獎號6': round(((data['獎號6']==N7).sum())/(6*(i-1))*100,2),
            '特別號': round(((data['特別號']==N8).sum())/(6*(i-1))*100,2),
            '頭獎數量': N9}, ignore_index=True)


# 建立一個ExcelWriter物件
writer = pd.ExcelWriter('大樂透開獎號碼-新版.xlsx', engine='openpyxl',mode='a')

# 將DataFrame寫入Excel
data_rate.to_excel(writer, sheet_name='機率', index=False)

writer.save()
