from selenium import webdriver
import time
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.load_workbook('C:\\Users\Arlen\Desktop\GitHub\win-the-lottery\大樂透開獎號碼-新版.xlsx')
ws = wb['工作表1']
k = ws.max_row

draw_number=ws.cell(k, 1).value
last_row_year=int(str(draw_number)[:3])
last_row_number=int(str(draw_number)[-3:])

driver = webdriver.Chrome("./goolemapSpider/chromedriver.exe")
driver.get('https://www.taiwanlottery.com.tw/lotto/Lotto649/history.aspx')
time.sleep(1)
date_format = "%y/%m/%d"
for draw_year in range(int(last_row_year),113):

    for i in range(int(last_row_number),120):


        flag = driver.find_element("name", 'Lotto649Control_history$txtNO')
        draw_number="000000" + str(i)
        flag.clear()
        time.sleep(0.7)

        '網業執行查詢'
        flag.send_keys(str(draw_year) + draw_number[-6:])
        time.sleep(0.7)
        driver.find_element("id",'Lotto649Control_history_btnSubmit').click()
        time.sleep(1)

        '判斷是否有內容'
        soup = BeautifulSoup(driver.page_source, "lxml")
        check=soup.find("span", {"id": "Lotto649Control_history_Label1"}).text

        if check=="查無資料":
            last_row_number=1
            break
        else:
            DrawTerm=soup.find("span", {"id": "Lotto649Control_history_dlQuery_L649_DrawTerm_0"}).text
            DDate=soup.find("span", {"id": "Lotto649Control_history_dlQuery_L649_DDate_0"}).text
            SellAmount=soup.find("span", {"id": "Lotto649Control_history_dlQuery_L649_SellAmount_0"}).text
            TotalAmount=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Total_0"}).text

            year, month, day = DDate.split("/")
            year = int(year) + 1911

            ws.cell(k,1,int(DrawTerm))
            ws.cell(k,2,"{}/{}/{}".format(year, month, day))
            ws.cell(k,3,int(SellAmount.replace(",", "")))
            ws.cell(k,4,int(TotalAmount.replace(",", "")))

            for i in range(1,7):
                locals()['SNo'+str(i)]=soup.find("span", {"id": "Lotto649Control_history_dlQuery_SNo" + str(i) +"_0"}).text
                ws.cell(k, 4+i,int(locals()['SNo'+str(i)]))

            SNo7=soup.find("span", {"id": "Lotto649Control_history_dlQuery_No7_0"}).text
            ws.cell(k, 11,int(SNo7))

            for i in 'ABC':
                locals()['Categ'+str(i)]=soup.find("span", {"id": "Lotto649Control_history_dlQuery_L649_Categ" + str(i) +"3_0"}).text
            ws.cell(k, 12,int(CategA.replace(",", "")))
            ws.cell(k, 13,int(CategB.replace(",", "")))
            ws.cell(k, 14,int(CategC.replace(",", "")))

            for i in range(2,7):
                locals()['label'+str(i)]=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label" + str(i) +"_0"}).text
                ws.cell(k, 13+i,int(locals()['label'+str(i)].replace(",", "")))

            categA_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_L649_CategA4_0"}).text
            categB_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label7_0"}).text
            categC_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label8_0"}).text
            label2_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label9_0"}).text
            label3_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label10_0"}).text
            label4_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label11_0"}).text
            label5_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label12_0"}).text
            label6_1=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label13_0"}).text

            categA_2=soup.find("span", {"id": "Lotto649Control_history_dlQuery_L649_CategA5_0"}).text
            categB_2=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label14_0"}).text
            categC_2=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label15_0"}).text
            label2_2=soup.find("span", {"id": "Lotto649Control_history_dlQuery_Label16_0"}).text

            ws.cell(k,20,int(categA_1.replace(",", "")))
            ws.cell(k,21,int(categB_1.replace(",", "")))
            ws.cell(k,22,int(categC_1.replace(",", "")))
            ws.cell(k,23,int(label2_1.replace(",", "")))
            ws.cell(k,24,int(label3_1.replace(",", "")))
            ws.cell(k,25,int(label4_1.replace(",", "")))
            ws.cell(k,26,int(label5_1.replace(",", "")))
            ws.cell(k,27,int(label6_1.replace(",", "")))
            ws.cell(k,28,int(categA_2.replace(",", "")))
            ws.cell(k,29,int(categB_2.replace(",", "")))
            ws.cell(k,30,int(categC_2.replace(",", "")))
            ws.cell(k,31,int(label2_2.replace(",", "")))

            print(DrawTerm,k)

            k += 1
            wb.save('大樂透開獎號碼-新版.xlsx')

